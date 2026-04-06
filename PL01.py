import streamlit as st
import pandas as pd
import io
import time
import os
import glob
import gc
import base64
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

# Lệnh set_page_config phải luôn nằm trên cùng
st.set_page_config(page_title="Phần mềm lập PL01 Chuyên nghiệp", page_icon="📊", layout="wide", initial_sidebar_state="expanded")

# ==========================================
# THUẬT TOÁN CHUẨN HÓA CHUỖI (FUZZY MATCHING)
# ==========================================
def normalize_text(text):
    """Gọt sạch tiếng Việt có dấu, khoảng trắng, ký tự đặc biệt để so khớp chuẩn xác."""
    if pd.isna(text): return ""
    t = str(text).lower().strip()
    t = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', t)
    t = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', t)
    t = re.sub(r'[ìíịỉĩ]', 'i', t)
    t = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', t)
    t = re.sub(r'[ùúụủũưừứựửữ]', 'u', t)
    t = re.sub(r'[ỳýỵỷỹ]', 'y', t)
    t = re.sub(r'[đ]', 'd', t)
    t = re.sub(r'[^a-z0-9]', '', t)
    return t

# ==========================================
# TÙY BIẾN GIAO DIỆN WEB (UI/UX DASHBOARD)
# ==========================================
st.markdown("""
<style>
    /* Font Inter hiện đại */
    @import url('https://fonts.googleapis.com/css2?family=Inter:opsz,wght@14..32,300;14..32,400;14..32,500;14..32,600;14..32,700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif !important;
    }

    /* Ẩn các thành phần mặc định của Streamlit tạo cảm giác App độc lập */
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}

    /* XÓA SẠCH CHỮ "Press Enter to apply" Ở MỌI Ô NHẬP LIỆU */
    div[data-testid="InputInstructions"] { display: none !important; }
    .stTextInput small { display: none !important; }
    div[data-baseweb="input"] + div { display: none !important; }

    /* Nền tổng thể Clean & Professional */
    .stApp {
        background-color: #f4f7f6;
    }

    /* Tiêu đề ứng dụng */
    .app-header {
        text-align: center;
        padding: 1.5rem 0 1rem 0;
        background: linear-gradient(90deg, #1e3a8a 0%, #2563eb 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        font-size: 2.4rem;
        font-weight: 800;
        letter-spacing: -0.5px;
        margin-bottom: 0.5rem;
    }
    .app-subheader {
        text-align: center;
        color: #64748b;
        font-size: 1.05rem;
        font-weight: 400;
        margin-bottom: 2rem;
    }

    /* TỰ ĐỘNG DÃN RỘNG SIDEBAR ĐỂ TÍCH CHỌN KHÔNG BỊ KÉO NGANG */
    section[data-testid="stSidebar"] {
        min-width: 430px !important;
        max-width: 500px !important;
        background-color: #ffffff;
        border-right: 1px solid #e2e8f0;
        box-shadow: 2px 0 8px rgba(0,0,0,0.02);
    }
    
    .company-name {
        text-align: center;
        font-size: 1.05rem;
        font-weight: 700;
        color: #0f172a;
        margin-bottom: 12px;
        text-transform: uppercase;
        letter-spacing: 0.3px;
        line-height: 1.4;
    }
    .station-name {
        text-align: center;
        color: #64748b;
        font-size: 0.85rem;
        font-weight: 600;
        margin-top: 5px;
        background: #f1f5f9;
        padding: 4px 8px;
        border-radius: 20px;
        display: inline-block;
        width: 100%;
    }
    .logo-container {
        display: flex;
        justify-content: center;
        margin-bottom: 15px;
    }
    .logo-img { width: 120px; height: auto; transition: transform 0.3s ease; }
    .logo-img:hover { transform: scale(1.03); }

    /* Làm nổi bật Ô nhập Mùa Vụ */
    div[data-baseweb="input"] {
        border: 1px solid #cbd5e1 !important;
        border-radius: 8px !important;
        background-color: #f8fafc !important;
    }
    div[data-baseweb="input"]:focus-within {
        border-color: #2563eb !important;
        box-shadow: 0 0 0 2px rgba(37, 99, 235, 0.2) !important;
        background-color: #ffffff !important;
    }
    
    /* Tùy biến Data Editor Ma trận ở Sidebar */
    [data-testid="stDataFrame"] {
        border-radius: 8px;
        border: 1px solid #e2e8f0;
        overflow: hidden;
    }

    /* Tabs Styling chuẩn Dashboard */
    div[data-testid="stTabs"] > div > div[data-testid="stTab"] {
        font-size: 1.05rem !important;
        font-weight: 600 !important;
        color: #64748b !important;
        padding: 0.75rem 1.5rem !important;
        background-color: transparent !important;
        border: none !important;
        border-bottom: 3px solid transparent !important;
    }
    div[data-testid="stTabs"] > div > div[data-testid="stTab"][aria-selected="true"] {
        color: #1e3a8a !important;
        border-bottom: 3px solid #2563eb !important;
    }
    div[data-testid="stTabs"] {
        background: white;
        padding: 10px 20px 20px 20px;
        border-radius: 12px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.03);
        border: 1px solid #e2e8f0;
    }

    /* Các Nút Bấm (Buttons) */
    .stButton > button {
        border-radius: 8px !important;
        font-weight: 600 !important;
        transition: all 0.2s ease !important;
        border: 1px solid #cbd5e1 !important;
        background: #ffffff !important;
        color: #334155 !important;
        box-shadow: 0 1px 2px rgba(0,0,0,0.05);
    }
    .stButton > button:hover {
        border-color: #94a3b8 !important;
        background: #f8fafc !important;
        color: #0f172a !important;
    }
    /* Primary Button (Nút Tích cực) */
    button[kind="primary"] {
        background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%) !important;
        color: white !important;
        border: none !important;
        box-shadow: 0 4px 6px -1px rgba(59, 130, 246, 0.3) !important;
    }
    button[kind="primary"]:hover {
        background: linear-gradient(135deg, #1e3a8a 0%, #2563eb 100%) !important;
        box-shadow: 0 6px 12px -2px rgba(59, 130, 246, 0.4) !important;
        transform: translateY(-1px);
    }

    /* Nút Download */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #059669 0%, #10b981 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        box-shadow: 0 4px 6px -1px rgba(16, 185, 129, 0.3) !important;
        width: 100%;
    }
    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #047857 0%, #059669 100%) !important;
        box-shadow: 0 6px 12px -2px rgba(16, 185, 129, 0.4) !important;
        transform: translateY(-1px);
    }

    /* File Uploader styling */
    [data-testid="stFileUploadDropzone"] {
        border: 2px dashed #94a3b8 !important;
        border-radius: 12px !important;
        background-color: #f8fafc !important;
        transition: all 0.3s;
    }
    [data-testid="stFileUploadDropzone"]:hover {
        border-color: #3b82f6 !important;
        background-color: #eff6ff !important;
    }

    /* Container Box mượt mà */
    [data-testid="stVerticalBlockBorderWrapper"] {
        border-radius: 12px;
        background-color: #ffffff;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -1px rgba(0, 0, 0, 0.03);
        border: 1px solid #e2e8f0;
        padding: 1rem;
    }
    
    /* Expander hiện đại */
    .streamlit-expanderHeader {
        font-weight: 600 !important;
        color: #1e293b !important;
        background-color: #f8fafc !important;
        border-radius: 8px !important;
    }
    div[data-testid="stExpander"] {
        border: 1px solid #e2e8f0 !important;
        border-radius: 8px !important;
        box-shadow: none !important;
        margin-bottom: 10px;
    }

    /* ===== HIỆU ỨNG TÁC GIẢ & LOGO ===== */
    .company-name {
        text-align: center;
        font-size: 1.1em;
        font-weight: 600;
        color: #1E3A8A;
        white-space: nowrap;
        overflow: hidden;
        animation: slideIn 1s ease-out, glow 2s ease-in-out infinite;
    }
    @keyframes slideIn {
        from { transform: translateX(-30px); opacity: 0; }
        to { transform: translateX(0); opacity: 1; }
    }
    @keyframes glow {
        0% { text-shadow: 0 0 0px #1E3A8A; }
        50% { text-shadow: 0 0 8px #3b82f6; }
        100% { text-shadow: 0 0 0px #1E3A8A; }
    }

    .station-name {
        text-align: center;
        color: gray;
        font-size: 0.9em;
        margin-top: 5px;
        animation: blinkSoft 1.5s step-start infinite;
    }
    @keyframes blinkSoft {
        0%, 100% { opacity: 1; color: #6c757d; }
        50% { opacity: 0.6; color: #1e3a8a; }
    }

    .logo-container {
        display: flex;
        justify-content: center;
        align-items: center;
        margin-bottom: 10px;
        transition: transform 0.3s ease;
    }
    .logo-container:hover { transform: scale(1.02); }
    .logo-img {
        width: 140px;
        height: auto;
        animation: gentlePulse 2s infinite;
    }
    @keyframes gentlePulse {
        0% { filter: drop-shadow(0 0 0px rgba(30,58,138,0.2)); transform: scale(1); }
        50% { filter: drop-shadow(0 0 6px rgba(30,58,138,0.4)); transform: scale(1.02); }
        100% { filter: drop-shadow(0 0 0px rgba(30,58,138,0.2)); transform: scale(1); }
    }

    .donate-box {
        position: relative;
        overflow: hidden;
        transition: all 0.3s ease;
        animation: borderGlow 2s infinite;
        background-color: #e6f3fd; 
        padding: 15px; 
        border-radius: 5px; 
        color: #0056b3; 
        font-size: 14.5px;
    }
    @keyframes borderGlow {
        0% { box-shadow: 0 0 0 0 rgba(0, 86, 179, 0.2); }
        50% { box-shadow: 0 0 0 8px rgba(0, 86, 179, 0.1); }
        100% { box-shadow: 0 0 0 0 rgba(0, 86, 179, 0.2); }
    }
    .marquee-text {
        display: inline-block;
        white-space: nowrap;
        animation: marquee 15s linear infinite;
        font-weight: 500;
    }
    @keyframes marquee {
        0% { transform: translateX(100%); }
        100% { transform: translateX(-100%); }
    }
    .blink-number {
        animation: blink 1s step-end infinite;
        font-size: 15px; 
        color: #d9534f;
    }
    @keyframes blink {
        0%, 100% { opacity: 1; }
        50% { opacity: 0.5; }
    }
</style>
""", unsafe_allow_html=True)

# ==========================================
# LỚP BẢO MẬT: KHÓA MẬT KHẨU (MÃ PIN) HIỆN ĐẠI
# ==========================================
def check_password():
    if "password_correct" not in st.session_state:
        st.session_state["password_correct"] = False

    if st.session_state["password_correct"]:
        return True

    # Đưa Form Đăng nhập lên cao hơn (giảm bớt <br>) và nới rộng cột giữa
    st.markdown("<br>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1.5, 1])
    with col2:
        with st.container(border=True):
            # Ép Tiêu đề nằm trên 1 dòng với white-space: nowrap và font-size vừa phải
            st.markdown("<h2 style='text-align: center; color: #1e3a8a; margin-bottom: 5px; font-size: 1.8rem; white-space: nowrap;'>🔐 ĐĂNG NHẬP HỆ THỐNG</h2>", unsafe_allow_html=True)
            st.markdown("<p style='text-align: center; color: #64748b; font-size: 0.95rem;'>Vui lòng nhập mật khẩu để sử dụng phần mềm.</p>", unsafe_allow_html=True)
            st.markdown("<hr style='border-top: 1px solid #e2e8f0; margin: 15px 0;'>", unsafe_allow_html=True)
            
            with st.form("login_form"):
                password = st.text_input("Mã PIN bảo mật", type="password", placeholder="Nhập Mật Khẩu...", label_visibility="collapsed")
                st.markdown("<br>", unsafe_allow_html=True)
                submitted = st.form_submit_button("🚀 Xác nhận Đăng nhập", use_container_width=True, type="primary")
                if submitted:
                    if password == "2685": 
                        st.session_state["password_correct"] = True
                        st.rerun()
                    else:
                        st.error("❌ Mật khẩu không chính xác.  \nLiên hệ Tác giả Trần Thọ: 098.7575.691")
    return False

if not check_password():
    st.stop()

# ==========================================
# GIAO DIỆN CHÍNH (MAIN APP HEADER)
# ==========================================
st.markdown("""
<style>
    .pm-quotes-container {
        background: linear-gradient(90deg, #1e3a8a 0%, #3b82f6 100%);
        color: white;
        padding: 12px 20px;
        border-radius: 8px;
        margin-bottom: 25px;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1);
        display: flex;
        align-items: center;
        overflow: hidden;
    }
    .pm-quotes-label {
        font-weight: 700;
        white-space: nowrap;
        margin-right: 15px;
        padding-right: 15px;
        border-right: 2px solid rgba(255,255,255,0.4);
        display: flex;
        align-items: center;
        gap: 8px;
    }
    .pm-marquee-wrapper {
        flex-grow: 1; 
        overflow: hidden; 
        white-space: nowrap;
    }
    .pm-marquee {
        display: inline-block;
        padding-left: 100%; 
        animation: marquee-quotes 30s linear infinite; 
        font-weight: 500;
        font-size: 0.95rem;
    }
    .pm-marquee:hover {
        animation-play-state: paused;
    }
    .pm-quote-text {
        font-style: italic;
    }
    .pm-quote-author {
        font-weight: 700;
        color: #fbbf24; 
    }
    @keyframes marquee-quotes {
        0% { transform: translateX(0); } 
        100% { transform: translateX(-100%); }
    }
</style>

<div class="pm-quotes-container">
    <div class="pm-quotes-label">
        <span>🇻🇳</span> Định hướng Chuyển đổi số
    </div>
    <div class="pm-marquee-wrapper">
        <span class="pm-marquee">
            <span class="pm-quote-text">"Chuyển đổi số là xu thế tất yếu, là đòi hỏi khách quan của sự phát triển..."</span> - <span class="pm-quote-author">Thủ tướng Chính phủ Phạm Minh Chính</span>
            &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;⭐&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
            <span class="pm-quote-text">"Dữ liệu là tài nguyên mới, là nền tảng của chuyển đổi số..."</span>
            &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;⭐&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
            <span class="pm-quote-text">"Chuyển đổi số phải lấy người dân, doanh nghiệp làm trung tâm, chủ thể, làm mục tiêu, động lực..."</span>
        </span>
    </div>
</div>
""", unsafe_allow_html=True)
# ==========================================
# 0. KHU VỰC CHÈN ẢNH CHỦ QUYỀN (SIDEBAR)
# ==========================================
script_dir = os.path.dirname(os.path.abspath(__file__))
image_files = glob.glob(os.path.join(script_dir, "anh_cua_toi*"))

st.sidebar.markdown("""
<div class="company-name" style="text-align: center; font-size: 1.1em; font-weight: 600; color: #1E3A8A; animation: glow 2s ease-in-out infinite;">
    Công ty TNHH MTV Khai Thác<br>Công Trình Thủy Lợi Kon Tum
</div>
""", unsafe_allow_html=True)

if image_files:
    with open(image_files[0], "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read()).decode()
    st.sidebar.markdown(
        f"""
        <div class="logo-container">
            <img src="data:image/png;base64,{encoded_string}" class="logo-img">
        </div>
        """,
        unsafe_allow_html=True
    )

# Thêm ô nhập liệu Tên Trạm ở Sidebar để người dùng tuỳ chỉnh
st.sidebar.markdown("<div style='font-size: 13px; font-weight: 600; color: #1e3a8a; margin-bottom: 5px;'>🏢 NHẬP TÊN TRẠM / ĐƠN VỊ CỦA BẠN</div>", unsafe_allow_html=True)
station_name = st.sidebar.text_input("Tên Trạm/Đơn vị:", value="TRẠM QLTN KHU VỰC I", label_visibility="collapsed")
st.sidebar.markdown(f'<div style="text-align: center;"><span class="station-name">✨ {station_name}</span></div>', unsafe_allow_html=True)
st.sidebar.markdown("<br>", unsafe_allow_html=True)

# ==========================================
# THIẾT LẬP MÙA VỤ THEO MA TRẬN NHÓM (GROUPED MATRIX)
# ==========================================
st.sidebar.markdown("### ⚙️ CẤU HÌNH MÙA VỤ")

st.sidebar.markdown("<div style='font-size: 13px; font-weight: 600; color: #1e3a8a; margin-bottom: 5px;'>📝 BƯỚC 1: Khai báo các Mùa Vụ trong năm</div>", unsafe_allow_html=True)
season_input_str = st.sidebar.text_input("Khai báo các Vụ", "Đông Xuân, Mùa", label_visibility="collapsed")
master_seasons = [s.strip() for s in season_input_str.split(',') if s.strip()]
if not master_seasons:
    master_seasons = ["Đông Xuân", "Mùa"] 

st.sidebar.markdown("<div style='font-size: 13px; font-weight: 600; color: #1e3a8a; margin-top: 15px; margin-bottom: 5px;'>🎯 BƯỚC 2: Chọn Mùa Vụ cho từng Nhóm</div>", unsafe_allow_html=True)
st.sidebar.markdown("<div style='font-size: 12.5px; color: #64748b; margin-bottom: 10px; font-style: italic;'>Mở các nhóm bên dưới và tích (☑) để gán Vụ.</div>", unsafe_allow_html=True)

# Từ điển ánh xạ ID cột sang Tên hiển thị rút gọn (Để bảng gọn gàng hơn)
COL_NAMES = {
    "9": "Chủ động (Cột 9)", "10": "CĐ 1 phần (10)", "11": "Tạo nguồn (11)",
    "12": "Chủ động (Cột 12)", "13": "CĐ 1 phần (13)", "14": "Tạo nguồn (14)",
    "16": "Chủ động (Cột 16)", "17": "CĐ 1 phần (17)", "18": "Tạo nguồn (18)",
    "19": "Chủ động (Cột 19)", "20": "CĐ 1 phần (20)", "21": "Tạo nguồn (21)",
    "23": "Chủ động (Cột 23)", "24": "CĐ 1 phần (24)", "25": "Tạo nguồn (25)",
    "26": "Chủ động (Cột 26)", "27": "CĐ 1 phần (27)", "28": "Tạo nguồn (28)",
    "29": "Thủy sản (Ao cá)"
}

# Hàm Render một bảng Ma trận nhỏ
def render_season_matrix(key_prefix, col_ids, master_seasons):
    df = pd.DataFrame({
        "Mã": col_ids,
        "Loại Hình": [COL_NAMES[c] for c in col_ids]
    })
    for s in master_seasons:
        df[s] = False # Mặc định bỏ trống

    col_config = {
        "Mã": None, # Ẩn cột ID đi cho đẹp
        "Loại Hình": st.column_config.TextColumn("Loại Hình", disabled=True)
    }
    for s in master_seasons:
        col_config[s] = st.column_config.CheckboxColumn(s, default=False)

    edited_df = st.data_editor(
        df,
        hide_index=True,
        column_config=col_config,
        use_container_width=True,
        key=f"editor_{key_prefix}"
    )
    
    local_cfg = {}
    for idx, row in edited_df.iterrows():
        col_id = str(row["Mã"])
        local_cfg[col_id] = [s for s in master_seasons if row[s]]
    return local_cfg

# Khởi tạo cfg trống và lần lượt gộp các nhóm vào
cfg = {}

with st.sidebar.expander("🌾 NHÓM LÚA", expanded=False):
    st.markdown("<div style='font-size:13.5px; font-weight:600; color:#334155; margin-bottom: 5px;'>💧 Tưới bằng Trọng lực</div>", unsafe_allow_html=True)
    cfg_lua_tl = render_season_matrix("lua_tl", ["9", "10", "11"], master_seasons)
    cfg.update(cfg_lua_tl)
    
    st.markdown("<div style='font-size:13.5px; font-weight:600; color:#334155; margin-top:15px; margin-bottom: 5px;'>🚰 Tưới bằng Động lực</div>", unsafe_allow_html=True)
    cfg_lua_dl = render_season_matrix("lua_dl", ["12", "13", "14"], master_seasons)
    cfg.update(cfg_lua_dl)

with st.sidebar.expander("🌳 NHÓM CÂY CÔNG NGHIỆP DÀI NGÀY", expanded=False):
    st.markdown("<div style='font-size:13.5px; font-weight:600; color:#334155; margin-bottom: 5px;'>💧 Tưới bằng Trọng lực</div>", unsafe_allow_html=True)
    cfg_cndn_tl = render_season_matrix("cndn_tl", ["16", "17", "18"], master_seasons)
    cfg.update(cfg_cndn_tl)
    
    st.markdown("<div style='font-size:13.5px; font-weight:600; color:#334155; margin-top:15px; margin-bottom: 5px;'>🚰 Tưới bằng Động lực</div>", unsafe_allow_html=True)
    cfg_cndn_dl = render_season_matrix("cndn_dl", ["19", "20", "21"], master_seasons)
    cfg.update(cfg_cndn_dl)

with st.sidebar.expander("🥬 NHÓM RAU, MÀU, CÂY CNNN", expanded=False):
    st.markdown("<div style='font-size:13.5px; font-weight:600; color:#334155; margin-bottom: 5px;'>💧 Tưới bằng Trọng lực</div>", unsafe_allow_html=True)
    cfg_cnnn_tl = render_season_matrix("cnnn_tl", ["23", "24", "25"], master_seasons)
    cfg.update(cfg_cnnn_tl)
    
    st.markdown("<div style='font-size:13.5px; font-weight:600; color:#334155; margin-top:15px; margin-bottom: 5px;'>🚰 Tưới bằng Động lực</div>", unsafe_allow_html=True)
    cfg_cnnn_dl = render_season_matrix("cnnn_dl", ["26", "27", "28"], master_seasons)
    cfg.update(cfg_cnnn_dl)

with st.sidebar.expander("🐟 NHÓM THỦY SẢN", expanded=False):
    st.markdown("<div style='font-size:13.5px; font-weight:600; color:#334155; margin-bottom: 5px;'>🌊 Nuôi trồng Thủy sản</div>", unsafe_allow_html=True)
    cfg_ts = render_season_matrix("ts", ["29"], master_seasons)
    cfg.update(cfg_ts)

# ==========================================
# KHÔI PHỤC NÚT DONATE CHỮ CHẠY GỐC
# ==========================================
st.sidebar.markdown("---")
st.sidebar.markdown("### ☕&nbsp;&nbsp;Góc nhỏ của Tác giả", unsafe_allow_html=True)

st.sidebar.markdown("""
<div class="donate-box">
<span class="marquee-text">Một ly cà phê từ bạn là sự ghi nhận tuyệt vời nhất cho những nỗ lực tự động hóa công việc này. Xin chân thành cảm ơn! ❤️</span>
<br><br>
🏦 <b>Ngân hàng:</b> Vietcom Bank<br>
💳 <b>STK:</b> <span class="blink-number"><b>0761002363642</b></span><br>
👤 <b>Chủ TK:</b> Trần Văn Thọ
</div>
""", unsafe_allow_html=True)

# ==========================================
# CÁC HÀM XỬ LÝ DATA & EXCEL
# ==========================================
COLS = [str(i) for i in range(1, 31)]

def to_float(val):
    try:
        if pd.isna(val) or str(val).strip() == "" or str(val) == "<NA>": return 0.0
        return float(str(val).replace(',', '').strip())
    except: return 0.0

def clean_zero(val):
    return val if val > 0 else ""

def clean_text(val):
    if pd.isna(val) or str(val) == "<NA>": return ""
    s = str(val).strip()
    if s.lower() in ['nan', 'none', 'nat', 'null', '<na>', '']: return ""
    if s.endswith('.0'): s = s[:-2]
    return s

def export_pl01_excel(df_raw, cfg, master_seasons, current_station_name):
    df_raw = df_raw.fillna("")
    wb = Workbook()
    ws = wb.active
    ws.title = "PL01"

    font_title = Font(name='Times New Roman', size=12, bold=True)
    font_bold = Font(name='Times New Roman', size=11, bold=True)
    font_italic = Font(name='Times New Roman', size=11, italic=True)
    font_normal = Font(name='Times New Roman', size=11)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 

    ws.append(["PHỤ LỤC 01. BẢNG KÊ ĐỐI TƯỢNG VÀ DIỆN TÍCH, BIỆN PHÁP TƯỚI, TIÊU ĐƯỢC HỖ TRỢ TIỀN SỬ DỤNG SẢN PHẨM, DỊCH VỤ CÔNG ÍCH THỦY LỢI GIAI ĐOẠN 2026-2030"] + [""]*29)
    ws.append(["TỔ CHỨC THỦY LỢI CƠ SỞ (HTXDVNN,......): (chỉ áp dụng đối với HTX Đoàn kết và các Công ty cà phê), XÃ/PHƯỜNG: ............................................................."] + [""]*29)
    
    ws.merge_cells('A1:AD1'); ws.merge_cells('A2:AD2')
    ws['A1'].font = font_title; ws['A1'].alignment = align_center
    ws['A2'].font = Font(name='Times New Roman', size=11, bold=True, italic=True); ws['A2'].alignment = align_center

    ws.append([
        "TT", "Hộ gia đình, cá nhân", "Bản đồ địa chính", "", "Diện tích thửa (m2)", 
        "Tên công trình cấp nước\n(Tuyến kênh, HCN, đập dâng,...)", "TỔNG DIỆN TÍCH (M2)", 
        "Tổng diện tích lúa", "DIỆN TÍCH TRỒNG LÚA (M2)"] + [""]*5 + 
        ["Tổng DT rau, màu, cây CNDN", "DIỆN TÍCH TRỒNG CÂY CÔNG NGHIỆP DÀI NGÀY(M2)"] + [""]*5 + 
        ["Tổng DT rau, màu, cây CNNN", "DIỆN TÍCH TRỒNG RAU, MÀU, CÂY CÔNG NGHIỆP NGẮN NGÀY (M2)"] + [""]*5 + 
        ["DIỆN TÍCH NUÔI TRỒNG THUỶ SẢN (M2)", "Ký xác nhận của đại diện hộ gia đình, cá nhân"]
    )
    ws.append([
        "", "", "Số tờ bản đồ", "Số thửa", "", "", "", "", 
        "Tưới tiêu bằng trọng lực", "", "", "Tưới tiêu bằng động lực", "", "", 
        "", "Tưới tiêu bằng trọng lực", "", "", "Tưới bằng động lực", "", "", 
        "", "Tưới tiêu bằng trọng lực", "", "", "Tưới bằng động lực", "", "", 
        "", ""
    ])
    ws.append([
        "", "", "", "", "", "", "", "", 
        "Chủ động", "CĐ 1 phần", "Tạo nguồn", "Chủ động", "CĐ 1 phần", "Tạo nguồn", 
        "", "Chủ động", "CĐ 1 phần", "Tạo nguồn", "Chủ động", "CĐ 1 phần", "Tạo nguồn", 
        "", "Chủ động", "CĐ 1 phần", "Tạo nguồn", "Chủ động", "CĐ 1 phần", "Tạo nguồn", 
        "", ""
    ])
    ws.append([
        "1", "2", "3", "4", "5", "6", "7=\n(8+15+22+29)", "8=\n(9+..+14)", 
        "9", "10", "11", "12", "13", "14", 
        "15=\n(16+..+21)", "16", "17", "18", "19", "20", "21", 
        "22=\n(23+..+28)", "23", "24", "25", "26", "27", "28", 
        "29", "30"
    ])

    merges = [
        'A3:A5', 'B3:B5', 'C3:D3', 'C4:C5', 'D4:D5', 'E3:E5', 'F3:F5', 'G3:G5', 'H3:H5',
        'I3:N3', 'I4:K4', 'L4:N4', 
        'O3:O5', 'P3:U3', 'P4:R4', 'S4:U4', 
        'V3:V5', 'W3:AB3', 'W4:Y4', 'Z4:AB4', 
        'AC3:AC5', 'AD3:AD5'
    ]
    for m in merges: ws.merge_cells(m)

    for r in range(3, 7):
        for c in range(1, 31):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align_center
            cell.border = thin_border
            if r == 6: cell.font = font_italic 
            else: cell.font = font_bold

    vertical_cols = [8, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19, 20, 22, 23, 24, 25, 26, 27, 28]

    all_selected_seasons = set()
    for s_list in cfg.values():
        all_selected_seasons.update(s_list)

    seasons = [s for s in master_seasons if s in all_selected_seasons]
    if not seasons: seasons = master_seasons

    num_seasons = len(seasons)
    current_excel_row = 7 

    def get_season_val(row, col_str, season_target):
        if season_target in cfg.get(col_str, []):
            return to_float(row.get(col_str))
        return 0.0

    projects = []
    for p in df_raw['6'].dropna().astype(str).str.strip():
        if p not in ["", "nan", "None", "<NA>"] and p not in projects:
            projects.append(p)

    if not projects:
        projects = ["(Trống)"]
        df_raw['6'] = "(Trống)"

    # Bộ nhớ đệm dành cho PL02 (Tổng hợp sang HA)
    pl02_totals = {p: {s: {c: 0.0 for c in range(6, 27)} for s in master_seasons} for p in projects}

    # XỬ LÝ THEO TỪNG BLOCK CÔNG TRÌNH (PROJECT) CHO PL01
    for project in projects:
        df_project = df_raw[df_raw['6'].astype(str).str.strip() == project]
        if df_project.empty: continue

        project_total_idx = current_excel_row
        project_season_idx = [current_excel_row + 1 + i for i in range(num_seasons)]
        data_start_row = current_excel_row + 1 + num_seasons

        data_rows = []
        ho_row_cursor = data_start_row
        tt = 1 

        for ho, group in df_project.groupby("2", sort=False, dropna=False):
            if pd.isna(ho) or str(ho).strip() == "" or str(ho) == "1": continue

            ho_parcels_exist = False
            ho_season_indices = []
            temp_ho_items = []
            ho_row_idx = ho_row_cursor

            for season_target in seasons:
                season_name = f"- Vụ {season_target}"
                season_parcels = []

                for _, row in group.iterrows():
                    l9 = get_season_val(row, "9", season_target)
                    l10 = get_season_val(row, "10", season_target)
                    l11 = get_season_val(row, "11", season_target)
                    l12 = get_season_val(row, "12", season_target)
                    l13 = get_season_val(row, "13", season_target)
                    l14 = get_season_val(row, "14", season_target)
                    
                    c16 = get_season_val(row, "16", season_target)
                    c17 = get_season_val(row, "17", season_target)
                    c18 = get_season_val(row, "18", season_target)
                    c19 = get_season_val(row, "19", season_target)
                    c20 = get_season_val(row, "20", season_target)
                    c21 = get_season_val(row, "21", season_target)
                    
                    m23 = get_season_val(row, "23", season_target)
                    m24 = get_season_val(row, "24", season_target)
                    m25 = get_season_val(row, "25", season_target)
                    m26 = get_season_val(row, "26", season_target)
                    m27 = get_season_val(row, "27", season_target)
                    m28 = get_season_val(row, "28", season_target)
                    
                    ca29 = get_season_val(row, "29", season_target)
                    
                    sum_total = sum([l9,l10,l11,l12,l13,l14, c16,c17,c18,c19,c20,c21, m23,m24,m25,m26,m27,m28, ca29])
                    if sum_total == 0: continue

                    # Lưu vào dữ liệu TỔNG HỢP PL02 (Chia 10000 để ra HA)
                    pl02_totals[project][season_target][6] += l9 / 10000.0
                    pl02_totals[project][season_target][7] += l10 / 10000.0
                    pl02_totals[project][season_target][8] += l11 / 10000.0
                    pl02_totals[project][season_target][9] += l12 / 10000.0
                    pl02_totals[project][season_target][10] += l13 / 10000.0
                    pl02_totals[project][season_target][11] += l14 / 10000.0
                    
                    pl02_totals[project][season_target][13] += c16 / 10000.0
                    pl02_totals[project][season_target][14] += c17 / 10000.0
                    pl02_totals[project][season_target][15] += c18 / 10000.0
                    pl02_totals[project][season_target][16] += c19 / 10000.0
                    pl02_totals[project][season_target][17] += c20 / 10000.0
                    pl02_totals[project][season_target][18] += c21 / 10000.0
                    
                    pl02_totals[project][season_target][20] += m23 / 10000.0
                    pl02_totals[project][season_target][21] += m24 / 10000.0
                    pl02_totals[project][season_target][22] += m25 / 10000.0
                    pl02_totals[project][season_target][23] += m26 / 10000.0
                    pl02_totals[project][season_target][24] += m27 / 10000.0
                    pl02_totals[project][season_target][25] += m28 / 10000.0
                    
                    pl02_totals[project][season_target][26] += ca29 / 10000.0

                    # Data dòng thửa cho PL01
                    r_data = [""] * 30
                    r_data[2] = clean_text(row.get("3"))
                    r_data[3] = clean_text(row.get("4"))
                    r_data[4] = to_float(row.get("5"))
                    r_data[5] = "" 
                    
                    r_data[8] = l9; r_data[9] = l10; r_data[10] = l11
                    r_data[11] = l12; r_data[12] = l13; r_data[13] = l14
                    
                    r_data[15] = c16; r_data[16] = c17; r_data[17] = c18
                    r_data[18] = c19; r_data[19] = c20; r_data[20] = c21
                    
                    r_data[22] = m23; r_data[23] = m24; r_data[24] = m25
                    r_data[25] = m26; r_data[26] = m27; r_data[27] = m28
                    
                    r_data[28] = ca29   
                    r_data[29] = clean_text(row.get("30"))
                    
                    season_parcels.append(r_data)

                if len(season_parcels) > 0:
                    ho_parcels_exist = True
                    season_row_idx = ho_row_idx + 1 + sum([len(s['parcels']) + 1 for s in temp_ho_items])
                    parcel_start = season_row_idx + 1
                    parcel_end = season_row_idx + len(season_parcels)

                    season_row_data = [""] * 30
                    season_row_data[1] = season_name
                    for i in vertical_cols:
                        col_letter = get_column_letter(i + 1)
                        if parcel_start == parcel_end: season_row_data[i] = f"={col_letter}{parcel_start}"
                        else: season_row_data[i] = f"=SUM({col_letter}{parcel_start}:{col_letter}{parcel_end})"

                    temp_ho_items.append({
                        "season_row_data": season_row_data,
                        "season_row_idx": season_row_idx,
                        "parcels": season_parcels
                    })
                    ho_season_indices.append(season_row_idx)

            if ho_parcels_exist:
                ho_row_data = [""] * 30
                ho_row_data[0] = tt
                ho_row_data[1] = ho
                for i in vertical_cols:
                    col_letter = get_column_letter(i + 1)
                    if len(ho_season_indices) == 1: ho_row_data[i] = f"={col_letter}{ho_season_indices[0]}"
                    else: ho_row_data[i] = f"=SUM({','.join([f'{col_letter}{idx}' for idx in ho_season_indices])})"

                data_rows.append({"type": "ho", "data": ho_row_data, "row_idx": ho_row_idx})
                ho_row_cursor += 1
                for s_item in temp_ho_items:
                    data_rows.append({"type": "season", "data": s_item["season_row_data"], "row_idx": s_item["season_row_idx"]})
                    ho_row_cursor += 1
                    for p_data in s_item["parcels"]:
                        data_rows.append({"type": "parcel", "data": p_data, "row_idx": ho_row_cursor})
                        ho_row_cursor += 1
                tt += 1

        project_max_row = ho_row_cursor - 1

        row_tong = ["1", "Tổng cộng", "", "", "", project] + [""] * 24
        row_tong[6] = f"={get_column_letter(8)}{project_total_idx}+{get_column_letter(15)}{project_total_idx}+{get_column_letter(22)}{project_total_idx}+{get_column_letter(29)}{project_total_idx}"
        row_tong[7] = f"=SUM({get_column_letter(9)}{project_total_idx}:{get_column_letter(14)}{project_total_idx})"
        row_tong[14] = f"=SUM({get_column_letter(16)}{project_total_idx}:{get_column_letter(21)}{project_total_idx})"
        row_tong[21] = f"=SUM({get_column_letter(23)}{project_total_idx}:{get_column_letter(28)}{project_total_idx})"

        if project_max_row >= data_start_row:
             for i in vertical_cols:
                 col_letter = get_column_letter(i + 1)
                 row_tong[i] = f"=SUM({col_letter}{project_season_idx[0]}:{col_letter}{project_season_idx[-1]})"

        ws.append([clean_zero(v) if isinstance(v, float) else v for v in row_tong])
        for c_idx, cell in enumerate(ws[project_total_idx], start=1):
            cell.font = Font(name='Times New Roman', size=11, bold=True, color="FF0000")
            cell.border = thin_border
            cell.alignment = align_center if c_idx != 2 else align_left
            cell.fill = fill_yellow 
            if c_idx >= 5 and c_idx <= 29: cell.number_format = '#,##0.00;-#,##0.00;""'

        alphabets = "abcdefghijklmnopqrstuvwxyz"
        for idx, s_name in enumerate(seasons):
            r_idx = project_season_idx[idx]
            r_data = [alphabets[idx % 26], f"Vụ {s_name}", "", "", "", ""] + [""] * 24
            r_data[6] = f"={get_column_letter(8)}{r_idx}+{get_column_letter(15)}{r_idx}+{get_column_letter(22)}{r_idx}+{get_column_letter(29)}{r_idx}"
            r_data[7] = f"=SUM({get_column_letter(9)}{r_idx}:{get_column_letter(14)}{r_idx})"
            r_data[14] = f"=SUM({get_column_letter(16)}{r_idx}:{get_column_letter(21)}{r_idx})"
            r_data[21] = f"=SUM({get_column_letter(23)}{r_idx}:{get_column_letter(28)}{r_idx})"

            if project_max_row >= data_start_row:
                for i in vertical_cols:
                    col_letter = get_column_letter(i + 1)
                    r_data[i] = f'=SUMIF($B${data_start_row}:$B${project_max_row}, "- Vụ {s_name}", {col_letter}${data_start_row}:{col_letter}${project_max_row})'

            ws.append([clean_zero(v) if isinstance(v, float) else v for v in r_data])
            for c_idx, cell in enumerate(ws[r_idx], start=1):
                cell.font = Font(name='Times New Roman', size=11, bold=True, color="FF0000")
                cell.border = thin_border
                cell.alignment = align_center if c_idx != 2 else align_left
                cell.fill = fill_yellow 
                if c_idx >= 5 and c_idx <= 29: cell.number_format = '#,##0.00;-#,##0.00;""'

        for item in data_rows:
            r_data = item["data"]
            row_idx = item["row_idx"]
            r_data[6] = f"=H{row_idx}+O{row_idx}+V{row_idx}+AC{row_idx}"
            r_data[7] = f"=SUM(I{row_idx}:N{row_idx})"
            r_data[14] = f"=SUM(P{row_idx}:U{row_idx})"
            r_data[21] = f"=SUM(W{row_idx}:AB{row_idx})"

            ws.append([clean_zero(v) if isinstance(v, float) else v for v in r_data])
            for col_idx, cell in enumerate(ws[row_idx], start=1):
                cell.border = thin_border
                cell.font = font_normal
                cell.alignment = align_center
                if col_idx == 2: cell.alignment = align_left
                if item["type"] == "ho" and col_idx in [1, 2]: cell.font = font_bold
                if col_idx >= 5 and col_idx <= 29: cell.number_format = '#,##0.00;-#,##0.00;""'

        current_excel_row = project_max_row + 1

    # ==========================================
    # CHÈN CHỮ KÝ / FOOTER CUỐI FILE PL01
    # ==========================================
    ws.append([""] * 30)
    ws.append([""] * 30)
    footer_start_row = ws.max_row + 1

    # Dòng Ngày tháng năm
    row_date = [""] * 30
    row_date[17] = "Ngày ..... tháng ..... năm ....."
    ws.append(row_date)
    ws.merge_cells(start_row=footer_start_row, start_column=18, end_row=footer_start_row, end_column=30)
    cell_date = ws.cell(row=footer_start_row, column=18)
    cell_date.font = Font(name='Times New Roman', size=11, italic=True)
    cell_date.alignment = Alignment(horizontal='center', vertical='center')

    # Dòng chức danh (Người lập, Trạm, Công ty)
    footer_row_2 = footer_start_row + 1
    row_sig = [""] * 30
    row_sig[0] = "Người lập"
    row_sig[6] = current_station_name.upper()
    row_sig[17] = "CÔNG TY TNHH MTV KHAI THÁC CTTL KON TUM"
    ws.append(row_sig)

    ws.merge_cells(start_row=footer_row_2, start_column=1, end_row=footer_row_2, end_column=6)
    ws.merge_cells(start_row=footer_row_2, start_column=7, end_row=footer_row_2, end_column=17)
    ws.merge_cells(start_row=footer_row_2, start_column=18, end_row=footer_row_2, end_column=30)

    for col in [1, 7, 18]:
        c = ws.cell(row=footer_row_2, column=col)
        c.font = Font(name='Times New Roman', size=11, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')

    # Chừa dòng trống ký tên Của Công Ty
    for _ in range(5):
        ws.append([""] * 30)

    # Khối chữ ký UBND Xã nằm ngay DƯỚI phần Công ty
    footer_ubnd_start = ws.max_row + 1
    
    # Dòng Phòng Kinh Tế (Trái) & UBND Xã (Phải)
    row_ubnd = [""] * 30
    row_ubnd[0] = "PHÒNG KINH TẾ"
    row_ubnd[17] = "UBND XÃ............." 
    ws.append(row_ubnd)
    
    # Dòng Ký tên đóng dấu
    row_ubnd_ky = [""] * 30
    row_ubnd_ky[17] = "(Ký tên, đóng dấu)" 
    ws.append(row_ubnd_ky)

    ws.merge_cells(start_row=footer_ubnd_start, start_column=1, end_row=footer_ubnd_start, end_column=6)
    ws.merge_cells(start_row=footer_ubnd_start, start_column=18, end_row=footer_ubnd_start, end_column=30)
    ws.merge_cells(start_row=footer_ubnd_start+1, start_column=18, end_row=footer_ubnd_start+1, end_column=30)

    c_pkt = ws.cell(row=footer_ubnd_start, column=1)
    c_pkt.font = Font(name='Times New Roman', size=11, bold=True, color="FF0000")
    c_pkt.alignment = Alignment(horizontal='center', vertical='center')

    c_ubnd = ws.cell(row=footer_ubnd_start, column=18)
    c_ubnd.font = Font(name='Times New Roman', size=11, bold=True)
    c_ubnd.alignment = Alignment(horizontal='center', vertical='center')
    
    c_ubnd_ky = ws.cell(row=footer_ubnd_start+1, column=18)
    c_ubnd_ky.font = Font(name='Times New Roman', size=11, italic=True)
    c_ubnd_ky.alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 8; ws.column_dimensions['D'].width = 8   
    for i in range(5, 31): ws.column_dimensions[get_column_letter(i)].width = 10


    # ==========================================
    # XÂY DỰNG SHEET PL02 (TỔNG HỢP DIỆN TÍCH - HA)
    # ==========================================
    ws2 = wb.create_sheet(title="PL02")

    ws2.append(["CÔNG TRÌNH THỦY LỢI DO CÔNG TY TNHH MTV KHAI THÁC CÔNG TRÌNH THỦY LỢI KON TUM QUẢN LÝ, KHAI THÁC"] + [""]*25)
    ws2.append(["PHỤ LỤC 02. TỔNG HỢP DIỆN TÍCH, BIỆN PHÁP TƯỚI, TIÊU ĐƯỢC HỖ TRỢ TIỀN SỬ DỤNG SẢN PHẨM, DỊCH VỤ CÔNG ÍCH THỦY LỢI GIAI ĐOẠN 2026-2030"] + [""]*25)
    ws2.append(["TRÊN ĐỊA BÀN: XÃ......................................................................"] + [""]*25)
    ws2.append(["TỈNH......................................................................"] + [""]*25)
    ws2.append([""]*26)

    for i in range(1, 5): ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=26)
    
    ws2['A1'].font = Font(name='Times New Roman', size=11)
    ws2['A1'].alignment = Alignment(horizontal="left", vertical="center")
    ws2['A2'].font = font_title; ws2['A2'].alignment = align_center
    ws2['A3'].font = font_bold; ws2['A3'].alignment = align_center
    ws2['A4'].font = font_bold; ws2['A4'].alignment = align_center

    # Bảng Tiêu đề PL02 (26 cột)
    ws2.append([
        "TT", "Tên trạm QLTN / Vụ", "Tên công trình cấp nước\n(Tuyến kênh, HCN, đập dâng.........)", 
        "TỔNG DIỆN TÍCH (HA)", "DIỆN TÍCH TRỒNG LÚA (HA)", "", "", "", "", "", "",
        "DIỆN TÍCH TRỒNG CÂY CÔNG NGHIỆP DÀI NGÀY (HA)", "", "", "", "", "", "",
        "DIỆN TÍCH TRỒNG RAU, MÀU, CÂY CÔNG NGHIỆP NGẮN NGÀY (HA)", "", "", "", "", "", "",
        "DIỆN TÍCH NUÔI TRỒNG THUỶ SẢN (HA)"
    ])
    
    ws2.append([
        "", "", "", "", 
        "Tổng diện tích lúa", "Tưới tiêu bằng trọng lực", "", "", "Tưới tiêu bằng động lực", "", "",
        "Tổng diện tích CNN dài ngày", "Tưới tiêu bằng trọng lực", "", "", "Tưới bằng động lực", "", "",
        "Tổng diện tích rau, màu, cây CNN ngắn ngày", "Tưới tiêu bằng trọng lực", "", "", "Tưới bằng động lực", "", "",
        ""
    ])
    
    ws2.append([
        "", "", "", "", 
        "", "Chủ động", "CĐ 1 phần", "Tạo nguồn", "Chủ động", "CĐ 1 phần", "Tạo nguồn",
        "", "Chủ động", "CĐ 1 phần", "Tạo nguồn", "Chủ động", "CĐ 1 phần", "Tạo nguồn",
        "", "Chủ động", "CĐ 1 phần", "Tạo nguồn", "Chủ động", "CĐ 1 phần", "Tạo nguồn",
        ""
    ])
    
    ws2.append([
        "1", "2", "3", "4=(5+12+19+26)", 
        "5=(6+..+11)", "6", "7", "8", "9", "10", "11",
        "12=(13+..+18)", "13", "14", "15", "16", "17", "18",
        "19=(20+..+25)", "20", "21", "22", "23", "24", "25",
        "26"
    ])

    merges2 = [
        'E6:K6', 'L6:R6', 'S6:Y6', 'Z6:Z8',
        'A6:A8', 'B6:B8', 'C6:C8', 'D6:D8',
        'E7:E8', 'L7:L8', 'S7:S8',
        'F7:H7', 'I7:K7', 'M7:O7', 'P7:R7', 'T7:V7', 'W7:Y7'
    ]
    for m in merges2: ws2.merge_cells(m)

    for r in range(6, 10):
        for c in range(1, 27):
            cell = ws2.cell(row=r, column=c)
            cell.alignment = align_center
            cell.border = thin_border
            if r == 9: cell.font = font_italic 
            else: cell.font = font_bold

    # Dữ liệu PL02
    # 1. Hàng TỔNG CỘNG
    ws2.append(["", "TỔNG CỘNG", ""] + [""] * 23)
    row_tong_idx = ws2.max_row

    # 2. Hàng Các Vụ
    season_total_rows = []
    for s in seasons:
        ws2.append(["", f"Vụ {s}", ""] + [""] * 23)
        season_total_rows.append((s, ws2.max_row))

    # 3. Hàng Trạm
    ws2.append(["I", current_station_name.upper(), ""] + [""] * 23)
    row_tram_idx = ws2.max_row

    # 4. Hàng Công trình
    project_totals_rows = []
    for i, project in enumerate(projects):
        if project == "(Trống)": continue
        ws2.append([i+1, "", project] + [""] * 23)
        proj_idx = ws2.max_row
        project_totals_rows.append(proj_idx)
        
        vu_indices_for_proj = []
        for s in seasons:
            r_vu = ["", f"- Vụ {s}", ""] + [""] * 23
            
            # Nạp dữ liệu HA đã được tổng hợp ở bước PL01
            for col_idx in range(6, 27):
                r_vu[col_idx-1] = pl02_totals[project][s][col_idx]
                
            ws2.append(r_vu)
            curr_idx = ws2.max_row
            vu_indices_for_proj.append(curr_idx)
            
            # Gắn công thức hàng ngang cho Vụ của Công trình
            ws2[f'D{curr_idx}'] = f"=E{curr_idx}+L{curr_idx}+S{curr_idx}+Z{curr_idx}"
            ws2[f'E{curr_idx}'] = f"=SUM(F{curr_idx}:K{curr_idx})"
            ws2[f'L{curr_idx}'] = f"=SUM(M{curr_idx}:R{curr_idx})"
            ws2[f'S{curr_idx}'] = f"=SUM(T{curr_idx}:Y{curr_idx})"
            
        # Gắn công thức hàng dọc và ngang cho dòng Tên Công Trình
        ws2[f'D{proj_idx}'] = f"=E{proj_idx}+L{proj_idx}+S{proj_idx}+Z{proj_idx}"
        ws2[f'E{proj_idx}'] = f"=SUM(F{proj_idx}:K{proj_idx})"
        ws2[f'L{proj_idx}'] = f"=SUM(M{proj_idx}:R{proj_idx})"
        ws2[f'S{proj_idx}'] = f"=SUM(T{proj_idx}:Y{proj_idx})"
        
        for col_l in ['F','G','H','I','J','K','M','N','O','P','Q','R','T','U','V','W','X','Y','Z']:
            if vu_indices_for_proj:
                ws2[f'{col_l}{proj_idx}'] = f"=SUM({col_l}{vu_indices_for_proj[0]}:{col_l}{vu_indices_for_proj[-1]})"
            else:
                ws2[f'{col_l}{proj_idx}'] = 0

    # 5. Gắn công thức cho dòng TRẠM
    for col_l in ['D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']:
        if project_totals_rows:
            ws2[f'{col_l}{row_tram_idx}'] = f"=SUM({','.join([f'{col_l}{idx}' for idx in project_totals_rows])})"
        else:
            ws2[f'{col_l}{row_tram_idx}'] = 0
            
    # 6. Gắn công thức cho dòng VỤ (Tổng)
    for s, s_idx in season_total_rows:
        for col_l in ['D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']:
            ws2[f'{col_l}{s_idx}'] = f'=SUMIF($B${row_tram_idx + 1}:$B${ws2.max_row}, "- Vụ {s}", {col_l}${row_tram_idx + 1}:{col_l}${ws2.max_row})'

    # 7. Gắn công thức cho dòng TỔNG CỘNG
    for col_l in ['D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']:
        ws2[f'{col_l}{row_tong_idx}'] = f"={col_l}{row_tram_idx}"

    # Căn chỉnh Border, Font, Nền Vàng cho Data PL02
    for r in range(10, ws2.max_row + 1):
        is_yellow = (r == row_tong_idx) or (r == row_tram_idx) or (r in [i for _, i in season_total_rows])
        for c in range(1, 27):
            cell = ws2.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = align_center if c != 2 and c != 3 else align_left
            if c >= 4: 
                cell.number_format = '#,##0.00;-#,##0.00;""'
            if is_yellow:
                cell.fill = fill_yellow
                cell.font = Font(name='Times New Roman', size=11, bold=True, color="FF0000")
            elif r in project_totals_rows:
                cell.font = font_bold
            else:
                cell.font = font_normal

    # Chèn Chữ ký Footer PL02
    ws2.append([""] * 26)
    ws2.append([""] * 26)
    f2_row = ws2.max_row + 1
    
    # Dòng 1: UBND Xã | TRẠM | Ngày tháng
    row_f2_1 = [""] * 26
    row_f2_1[1] = "UBND XÃ..............."
    row_f2_1[9] = current_station_name.upper()
    row_f2_1[18] = "Ngày ..... tháng ..... năm ....."
    ws2.append(row_f2_1)
    
    # Dòng 2: Ký tên | | Công ty
    row_f2_2 = [""] * 26
    row_f2_2[1] = "(Ký tên, đóng dấu)"
    row_f2_2[18] = "CÔNG TY TNHH MTV KHAI THÁC CTTL KON TUM"
    ws2.append(row_f2_2)
    
    # Các dòng khoảng trống Ký
    for _ in range(4): ws2.append([""] * 26)
    
    # Dòng Ký tên cho Trạm và Công Ty (ở dưới cùng)
    row_f2_last = [""] * 26
    row_f2_last[9] = "(Ký tên, đóng dấu)"
    row_f2_last[18] = "(Ký tên, đóng dấu)"
    ws2.append(row_f2_last)

    # Merge và Style cho Footer PL02
    # Cột 1: UBND (Index B)
    ws2.merge_cells(start_row=f2_row, start_column=2, end_row=f2_row, end_column=6)
    ws2.merge_cells(start_row=f2_row+1, start_column=2, end_row=f2_row+1, end_column=6)
    
    # Cột 2: TRẠM (Index J)
    ws2.merge_cells(start_row=f2_row, start_column=10, end_row=f2_row, end_column=14)
    ws2.merge_cells(start_row=ws2.max_row, start_column=10, end_row=ws2.max_row, end_column=14)
    
    # Cột 3: CÔNG TY (Index S)
    ws2.merge_cells(start_row=f2_row, start_column=19, end_row=f2_row, end_column=26)
    ws2.merge_cells(start_row=f2_row+1, start_column=19, end_row=f2_row+1, end_column=26)
    ws2.merge_cells(start_row=ws2.max_row, start_column=19, end_row=ws2.max_row, end_column=26)

    # Style
    for c in [2, 10, 19]:
        cell1 = ws2.cell(row=f2_row, column=c)
        cell1.alignment = Alignment(horizontal='center', vertical='center')
        cell2 = ws2.cell(row=f2_row+1, column=c)
        cell2.alignment = Alignment(horizontal='center', vertical='center')
        cell3 = ws2.cell(row=ws2.max_row, column=c)
        cell3.alignment = Alignment(horizontal='center', vertical='center')
        
        if c == 19: # Dòng Ngày tháng in nghiêng
            cell1.font = Font(name='Times New Roman', size=11, italic=True)
            cell2.font = Font(name='Times New Roman', size=11, bold=True)
            cell3.font = Font(name='Times New Roman', size=11, italic=True)
        elif c == 10: # Trạm
            cell1.font = Font(name='Times New Roman', size=11, bold=True)
            cell3.font = Font(name='Times New Roman', size=11, italic=True)
        else: # UBND
            cell1.font = Font(name='Times New Roman', size=11, bold=True)
            cell2.font = Font(name='Times New Roman', size=11, italic=True)

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 30
    for i in range(4, 27): ws2.column_dimensions[get_column_letter(i)].width = 11

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def export_formatted_data_goc(df):
    df_clean = df.fillna("")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data_Goc"

    font_bold = Font(name='Times New Roman', size=11, bold=True)
    font_normal = Font(name='Times New Roman', size=11)
    font_italic = Font(name='Times New Roman', size=11, italic=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(["PHỤ LỤC 01. BẢNG KÊ ĐỐI TƯỢNG VÀ DIỆN TÍCH... (DATA NỘI BỘ ĐỂ NHẬP LIỆU)"] + [""]*29)
    ws.append(["TỔ CHỨC THỦY LỢI CƠ SỞ..."] + [""]*29)
    ws.append([
        "TT", "Hộ gia đình, cá nhân", "Bản đồ địa chính", "", "Diện tích thửa (m2)", "Tên công trình cấp nước", "TỔNG DIỆN TÍCH (M2)", "Tổng diện tích lúa", "DIỆN TÍCH TRỒNG LÚA (M2)"] + [""]*5 + ["Tổng DT rau, màu, cây CNDN", "DIỆN TÍCH TRỒNG CÂY CÔNG NGHIỆP DÀI NGÀY(M2)"] + [""]*5 + ["Tổng DT rau, màu, cây CNNN", "DIỆN TÍCH TRỒNG RAU, MÀU, CÂY CÔNG NGHIỆP NGẮN NGÀY (M2)"] + [""]*5 + ["DIỆN TÍCH NUÔI TRỒNG THUỶ SẢN (M2)", "Ký xác nhận"]
    )
    ws.append(["", "", "Số tờ bản đồ", "Số thửa", "", "", "", "", "Tưới tiêu bằng trọng lực", "", "", "Tưới tiêu bằng động lực", "", "", "", "Tưới tiêu bằng trọng lực", "", "", "Tưới bằng động lực", "", "", "", "Tưới tiêu bằng trọng lực", "", "", "Tưới bằng động lực", "", "", "", ""])
    ws.append(["", "", "", "", "", "", "", "", "Chủ động", "CĐ 1 phần", "Tạo nguồn", "Chủ động", "CĐ 1 phần", "Tạo nguồn", "", "Chủ động", "CĐ 1 phần", "Tạo nguồn", "Chủ động", "CĐ 1 phần", "Tạo nguồn", "", "Chủ động", "CĐ 1 phần", "Tạo nguồn", "Chủ động", "CĐ 1 phần", "Tạo nguồn", "", ""])
    ws.append(COLS)

    merges = ['A3:A5', 'B3:B5', 'C3:D3', 'C4:C5', 'D4:D5', 'E3:E5', 'F3:F5', 'G3:G5', 'H3:H5', 'I3:N3', 'I4:K4', 'L4:N4', 'O3:O5', 'P3:U3', 'P4:R4', 'S4:U4', 'V3:V5', 'W3:AB3', 'W4:Y4', 'Z4:AB4', 'AC3:AC5', 'AD3:AD5']
    for m in merges: ws.merge_cells(m)

    for r in range(3, 7):
        for c in range(1, 31):
            cell = ws.cell(row=r, column=c)
            cell.alignment = align_center
            cell.border = thin_border
            cell.font = font_bold if r < 6 else font_italic
            if r == 6: cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Nhận diện danh sách dự án
    projects = []
    if '6' in df_clean.columns:
        for p in df_clean['6'].astype(str).str.strip():
            if p not in ["", "nan", "None", "<NA>", "0", "0.0"] and p not in projects:
                projects.append(p)
                
    if not projects:
        projects = ["(Trống)"]

    # Xử lý gộp dòng cho file DATA NỘI BỘ (Chừa 1 dòng trống điền tên CT)
    for project in projects:
        df_project = df_clean[df_clean['6'].astype(str).str.strip() == project] if '6' in df_clean.columns else df_clean
        if df_project.empty: continue
        
        # 1. Hàng Header chỉ chứa tên Công trình (Bôi vàng giống ý người dùng)
        proj_row = [""] * 30
        proj_row[5] = project
        ws.append(proj_row)
        header_row_idx = ws.max_row
        for c in range(1, 31):
            cell = ws.cell(row=header_row_idx, column=c)
            cell.border = thin_border
            if c == 6:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell.font = Font(name='Times New Roman', size=11, bold=True)
                cell.alignment = align_center

        # 2. Các hàng Dữ liệu thửa đất (Xóa ẩn tên công trình đi)
        for _, row in df_project.iterrows():
            row_data = list(row)
            row_data[5] = "" # Chừa trống Tên CT
            
            clean_row = []
            for i, v in enumerate(row_data):
                if i >= 4 and i <= 28 and i != 5: # Fomat các cột số liệu diện tích
                    val = to_float(v)
                    clean_row.append(clean_zero(val))
                else:
                    clean_row.append(v)

            ws.append(clean_row)
            
            data_row_idx = ws.max_row
            for col_idx, cell in enumerate(ws[data_row_idx], start=1):
                cell.font = font_normal
                cell.border = thin_border
                cell.alignment = align_left if col_idx == 2 else align_center
                
                if col_idx >= 5 and col_idx <= 29:
                    cell.number_format = '#,##0.00;-#,##0.00;""'

    ws.column_dimensions['B'].width = 25
    for i in range(5, 31): ws.column_dimensions[get_column_letter(i)].width = 11

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ==========================================
# MAIN TABS (CHUYỂN ĐỔI CHỨC NĂNG)
# ==========================================
tab1, tab2 = st.tabs(["📊 XÂY DỰNG BÁO CÁO PL01 & PL02", "🕵️ KIỂM TRA PL01 & PHỤC HỒI DATA"])

# ------------------------------------------
# TAB 1: XÂY DỰNG BÁO CÁO PL01
# ------------------------------------------
with tab1:
    st.markdown("<h3 style='color: #1e293b; font-size: 1.3rem; margin-top: 10px;'>📥 Tải Dữ liệu Đầu vào</h3>", unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader("Upload file Data Excel", type=["xlsx", "xls"], label_visibility="collapsed")

    if uploaded_file is not None:
        try:
            # Đọc danh sách sheet
            xls = pd.ExcelFile(uploaded_file)
            sheet_names = xls.sheet_names
            
            if len(sheet_names) > 1:
                selected_sheet = st.selectbox("📂 File có nhiều Sheet. Vui lòng chọn Sheet chứa Data Gốc:", sheet_names)
            else:
                selected_sheet = sheet_names[0]
                
            df_check_type = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=None, nrows=15)
            is_pl01_file = False
            for i, row in df_check_type.iterrows():
                vals = [str(x).strip() for x in row.values]
                if any("tổng cộng" in val.lower() for val in vals):
                    is_pl01_file = True
                    break
            
            if is_pl01_file:
                st.error("🚫 **CẢNH BÁO:** Bạn đang chọn Sheet chứa Báo cáo **PL01** vào khu vực Tạo báo cáo! Vui lòng chỉ chọn Sheet **Data Gốc** vào đây. (Nếu muốn kiểm tra file PL01, hãy chuyển sang Tab 2).")
                st.session_state.pop('raw_data', None)
            else:
                if st.session_state.get('last_file_id') != f"{uploaded_file.file_id}_{selected_sheet}":
                    st.session_state['last_file_id'] = f"{uploaded_file.file_id}_{selected_sheet}"
                    
                    for key in ['raw_data', 'pl01_data', 'goc_data', 'cfg_hash']:
                        if key in st.session_state:
                            del st.session_state[key]
                    gc.collect()
                    
                    progress_text = "⏳ Đang phân tích và đồng bộ hóa cấu trúc file..."
                    my_bar = st.progress(0, text=progress_text)
                    for percent in range(100):
                        time.sleep(0.005)
                        my_bar.progress(percent + 1, text=f"{progress_text} {percent + 1}%")
                    my_bar.empty()
                    
                    # Đọc đúng sheet đã chọn
                    df_raw = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=None)
                    
                    col_map = {}
                    header_idx = -1
                    
                    for i, row in df_raw.head(20).iterrows():
                        vals = [str(x).strip().replace('.0', '') for x in row.values]
                        if '1' in vals and '2' in vals and '3' in vals and '9' in vals:
                            header_idx = i
                            for c_idx, val in enumerate(vals):
                                if val in COLS:
                                    col_map[val] = c_idx
                            break
                            
                    if header_idx == -1 or not col_map:
                        temp_header = [""] * df_raw.shape[1]
                        for r in range(min(10, len(df_raw))):
                            row_vals = df_raw.iloc[r].values
                            current_val = ""
                            for c in range(len(row_vals)):
                                val = str(row_vals[c]).strip()
                                if val and val.lower() != 'nan':
                                    current_val = val
                                else:
                                    val = current_val 
                                temp_header[c] += " " + normalize_text(val)
                        
                        for c_idx, text in enumerate(temp_header):
                            if "hogiadinh" in text or "canhan" in text: col_map['2'] = c_idx
                            elif "bando" in text or "soto" in text: col_map['3'] = c_idx
                            elif "sothua" in text: col_map['4'] = c_idx
                            elif "dientichthua" in text: col_map['5'] = c_idx
                            elif "tencongtrinh" in text: col_map['6'] = c_idx
                            elif "luachudong" in text and "trongluc" in text: col_map['9'] = c_idx
                        
                        for i, row in df_raw.head(20).iterrows():
                            vals = [str(x).strip().replace('.0', '') for x in row.values]
                            if 'Tổng cộng' in vals or any("Vụ" in v for v in vals):
                                header_idx = max(0, i - 1)
                                break
                            
                    if header_idx != -1 and col_map:
                        data_part = df_raw.iloc[header_idx+1:].reset_index(drop=True)
                        extracted_rows = []
                        current_ho = ""
                        current_cong_trinh = "" 
                        
                        for _, row in data_part.iterrows():
                            c2_idx = col_map.get('2', -1)
                            c3_idx = col_map.get('3', -1)
                            c4_idx = col_map.get('4', -1)
                            c6_idx = col_map.get('6', -1)
                            
                            c2 = str(row.iloc[c2_idx]).strip() if c2_idx != -1 and pd.notna(row.iloc[c2_idx]) else ""
                            c3 = str(row.iloc[c3_idx]).strip() if c3_idx != -1 and pd.notna(row.iloc[c3_idx]) else ""
                            c4 = str(row.iloc[c4_idx]).strip() if c4_idx != -1 and pd.notna(row.iloc[c4_idx]) else ""
                            c6 = str(row.iloc[c6_idx]).strip() if c6_idx != -1 and pd.notna(row.iloc[c6_idx]) else ""
                            
                            if c2.startswith('Tổng cộng') or (c2.startswith('Vụ ') and not c2.startswith('- Vụ ')):
                                continue
                                
                            if c2 != "" and c3 in ["", "nan", "None", "<NA>"] and c4 in ["", "nan", "None", "<NA>"] and not c2.startswith("- Vụ"):
                                current_ho = c2
                                continue
                            
                            if c6 != "" and c2 in ["", "nan", "None", "<NA>"] and c3 in ["", "nan", "None", "<NA>"]:
                                current_cong_trinh = c6
                                continue
                                
                            if (c3 not in ["", "nan", "None", "<NA>"]) or (c4 not in ["", "nan", "None", "<NA>"]):
                                actual_ho = c2 if (c2 not in ["", "nan", "None", "<NA>"] and not c2.startswith("- Vụ")) else current_ho
                                
                                new_row_data = [""] * 30
                                for col_num in COLS:
                                    c_idx = col_map.get(col_num, -1)
                                    if c_idx != -1 and c_idx < len(row):
                                        new_row_data[int(col_num)-1] = row.iloc[c_idx]
                                
                                new_row = pd.Series(new_row_data, index=COLS)
                                new_row['2'] = actual_ho
                                
                                if str(new_row['6']).strip() in ["", "nan", "None", "<NA>"] and current_cong_trinh != "":
                                    new_row['6'] = current_cong_trinh
                                    
                                new_row['7'] = 0; new_row['8'] = 0; new_row['15'] = 0; new_row['22'] = 0
                                extracted_rows.append(new_row)
                                
                        df_final = pd.DataFrame(extracted_rows, columns=COLS)
                        df_final = df_final.dropna(subset=['2'])
                        df_final = df_final[df_final['2'].astype(str).str.strip() != ""]
                        for col in ['3', '4']: df_final[col] = df_final[col].apply(clean_text)
                        
                        st.session_state.raw_data = df_final
                        st.session_state.pop('pl01_data', None)
                        st.session_state.pop('goc_data', None)
                        st.toast("✅ Đã nhận diện và bóc tách dữ liệu thành công!", icon="🚀")
                    else:
                        st.error("❌ File không đúng định dạng. Không tìm thấy Hàng chứa số thứ tự cột (1 -> 30) hoặc các Tiêu đề cốt lõi.")

        except Exception as e:
            st.error(f"❌ Lỗi đọc file: Vui lòng kiểm tra lại định dạng file Excel. Lỗi chi tiết: {e}")

    if 'raw_data' in st.session_state:
        st.markdown("<hr style='margin: 20px 0;'>", unsafe_allow_html=True)
        st.markdown("<h3 style='color: #1e293b; font-size: 1.3rem; margin-bottom: 15px;'>🔎 Radar Quét Lỗi Dữ Liệu</h3>", unsafe_allow_html=True)
        
        with st.container(border=True):
            df_check = st.session_state.raw_data.copy()
            df_check['2'] = df_check['2'].astype(str).str.strip()
            df_check['3'] = df_check['3'].astype(str).str.strip()
            df_check['4'] = df_check['4'].astype(str).str.strip()
            
            valid_dup_mask = (df_check['3'] != '') & (df_check['4'] != '') & (df_check['3'] != 'nan') & (df_check['4'] != 'nan') & (df_check['3'] != '<NA>') & (df_check['4'] != '<NA>')
            df_valid_dup = df_check[valid_dup_mask]
            
            dup_mask = df_valid_dup.duplicated(subset=['3', '4'], keep=False)
            if dup_mask.any():
                dup_df = df_valid_dup[dup_mask]
                dup_summary = dup_df.groupby(['3', '4']).size().reset_index(name='count')
                st.error("⚠️ PHÁT HIỆN TRÙNG LẶP THỬA ĐẤT (Hệ thống phát hiện các thửa đất đang bị tách thành nhiều dòng)")
                for _, r in dup_summary.iterrows():
                    st.write(f"👉 **Tờ bản đồ {r['3']}, thửa {r['4']}**: xuất hiện **{r['count']}** lần.")
                
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    if st.button("🗑️ Xóa tất cả thửa trùng (Giữ lại 1)"):
                        dup_pairs = dup_summary[['3','4']].values.tolist()
                        mask = pd.Series(False, index=st.session_state.raw_data.index)
                        for to, thua in dup_pairs:
                            idxs = st.session_state.raw_data[(st.session_state.raw_data['3'] == to) & (st.session_state.raw_data['4'] == thua)].index
                            if len(idxs) > 1:
                                mask[idxs[1:]] = True
                        st.session_state.raw_data = st.session_state.raw_data[~mask].copy()
                        st.success(f"✅ Đã xóa {mask.sum()} dòng trùng (giữ lại 1 dòng cho mỗi thửa).")
                        st.rerun()
            else:
                st.success("✅ Kiểm tra Trùng lặp: Không phát hiện lỗi trùng lặp thửa đất.")

            invalid_area_rows = []
            crop_cols = ['9','10','11','12','13','14', '16','17','18','19','20','21', '23','24','25','26','27','28', '29']
            
            for col in crop_cols:
                df_check[col] = df_check[col].apply(to_float)
            df_check['5'] = df_check['5'].apply(to_float)
            df_check['tong_hotro_dong'] = df_check[crop_cols].sum(axis=1)

            df_check['group_id'] = df_check.apply(
                lambda r: f"{r['2']}_{r['3']}_{r['4']}" if (r['3'] not in ['', 'nan', '<NA>'] and r['4'] not in ['', 'nan', '<NA>']) else f"row_{r.name}",
                axis=1
            )

            for name, group in df_check.groupby('group_id'):
                dt_thua = group['5'].max() 
                tong_hotro_cua_thua = group['tong_hotro_dong'].sum() 

                if round(tong_hotro_cua_thua, 2) > round(dt_thua, 2):
                    first_row = group.iloc[0]
                    invalid_area_rows.append({
                        'ho': first_row['2'],
                        'to_bd': first_row['3'] if first_row['3'] not in ['', 'nan', '<NA>'] else '(Trống)',
                        'thua': first_row['4'] if first_row['4'] not in ['', 'nan', '<NA>'] else '(Trống)',
                        'dt_thua': dt_thua,
                        'dt_hotro': tong_hotro_cua_thua
                    })

            if invalid_area_rows:
                st.error("🚨 LỖI LOGIC: TỔNG DIỆN TÍCH TƯỚI VƯỢT QUÁ DIỆN TÍCH ĐỊA CHÍNH")
                for r in invalid_area_rows:
                    st.write(f"👉 **Chủ hộ: {r['ho']} | Tờ {r['to_bd']}, thửa {r['thua']}**: DT Thửa = **{r['dt_thua']:,.2f}** < Tổng Tưới = **{r['dt_hotro']:,.2f}**")
            else:
                st.success("✅ Kiểm tra Diện tích: Hoàn toàn hợp lệ, không có thửa nào vượt quá hạn mức.")

        st.markdown("<hr style='margin: 20px 0;'>", unsafe_allow_html=True)
        st.markdown("<h3 style='color: #1e293b; font-size: 1.3rem; margin-bottom: 10px;'>🚀 Xuất Biểu mẫu Báo cáo</h3>", unsafe_allow_html=True)
        
        if st.button("TỔNG HỢP VÀ TẠO BÁO CÁO (EXCEL)", type="primary", use_container_width=True):
            progress_text = "⏳ Đang tính toán Ma trận và Gắn CÔNG THỨC EXCEL..."
            my_bar = st.progress(0, text=progress_text)
            for percent in range(1, 40):
                time.sleep(0.005)
                my_bar.progress(percent, text=f"{progress_text} {percent}%")
                
            my_bar.progress(40, text="⏳ HỆ THỐNG ĐANG CHẠY...")
            pl01_data = export_pl01_excel(st.session_state.raw_data, cfg, master_seasons, station_name)
            
            my_bar.progress(70, text="⏳ VUI LÒNG CHỜ TRONG GIÂY LÁT... ")
            goc_data = export_formatted_data_goc(st.session_state.raw_data)
            
            for percent in range(70, 101):
                time.sleep(0.005)
                my_bar.progress(percent, text=f"🎉 ĐÃ HOÀN THÀNH XONG! {percent}%")
            time.sleep(0.5)
            my_bar.empty()
            
            st.session_state['pl01_data'] = pl01_data
            st.session_state['goc_data'] = goc_data
            st.session_state['cfg_hash'] = str(cfg) + str(master_seasons) + station_name

        if 'pl01_data' in st.session_state:
            if st.session_state.get('cfg_hash') == str(cfg) + str(master_seasons) + station_name:
                st.markdown("<br>", unsafe_allow_html=True)
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(label="📥 TẢI XUỐNG BÁO CÁO PL01 & PL02", data=st.session_state['pl01_data'], file_name="BieuMau_PL01_PL02.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                with col2:
                    st.download_button(label="🔄 TẢI FILE DATA NỘI BỘ", data=st.session_state['goc_data'], file_name="Data_Goc.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.warning("⚠️ Cấu hình mùa vụ hoặc Tên Trạm đã thay đổi. Vui lòng bấm 'TỔNG HỢP VÀ TẠO BÁO CÁO' lại để cập nhật.")

# ------------------------------------------
# TAB 2: KIỂM TRA & PHỤC HỒI DATA TỪ PL01
# ------------------------------------------
with tab2:
    st.markdown("<h3 style='color: #1e293b; font-size: 1.3rem; margin-top: 10px;'>📥 Tải file PL01 Cần Kiểm tra / Phục hồi</h3>", unsafe_allow_html=True)
    st.info("Hệ thống sẽ quét file PL01 để phát hiện các thửa đất nhập trùng trong cùng Vụ, đồng thời trích xuất ngược lại dữ liệu Data Gốc thông minh.")

    check_file = st.file_uploader("Upload file PL01", type=["xlsx", "xls"], key="check_file", label_visibility="collapsed")

    if check_file is not None:
        try:
            # Chọn Sheet cho Tab 2
            xls_check = pd.ExcelFile(check_file)
            sheet_names_check = xls_check.sheet_names
            
            if len(sheet_names_check) > 1:
                selected_sheet_check = st.selectbox("📂 File có nhiều Sheet. Vui lòng chọn Sheet chứa Báo Cáo PL01:", sheet_names_check)
            else:
                selected_sheet_check = sheet_names_check[0]
                
            df_check = pd.read_excel(check_file, sheet_name=selected_sheet_check, header=None)
            
            col_map_pl01 = {}
            start_row_idx = -1
            
            # Tối ưu hóa việc tìm Start Row (Bắt ngay sau dòng đánh số 1, 2, 3...)
            for i, row in df_check.head(20).iterrows():
                vals = [str(x).strip().replace('.0', '') for x in row.values]
                if '1' in vals and '2' in vals and '3' in vals:
                    for c_idx, val in enumerate(vals):
                        if val in COLS:
                            col_map_pl01[val] = c_idx
                    start_row_idx = i + 1 
                    break
                    
            if start_row_idx == -1:
                st.error("❌ Không tìm thấy cấu trúc chuẩn. Vui lòng đảm bảo đây là file PL01 (có chứa hàng đánh số cột từ 1 đến 30).")
            else:
                st.markdown("<br><h3 style='color: #1e293b; font-size: 1.2rem; margin-bottom: 10px;'>🔎 Phân tích Dữ liệu File tải lên</h3>", unsafe_allow_html=True)
                with st.container(border=True):
                    data = df_check.iloc[start_row_idx:].copy()
                    current_season = "Không xác định"
                    parcels = []
                    
                    idx_1 = col_map_pl01.get('1', 0)
                    idx_2 = col_map_pl01.get('2', 1)
                    idx_3 = col_map_pl01.get('3', 2)
                    idx_4 = col_map_pl01.get('4', 3)
                    
                    for idx, row in data.iterrows():
                        # Dừng Radar nếu chạm chữ ký
                        row_str = " ".join([str(x).lower() for x in row.values if pd.notna(x)])
                        if "người lập" in row_str or "phòng kinh tế" in row_str or "ký tên" in row_str or "ngày" in str(row.iloc[17] if len(row)>17 else "").lower():
                            break
                            
                        excel_row_num = idx + 1 
                        
                        col2_name = str(row[idx_2]).strip() if idx_2 < len(row) and pd.notna(row[idx_2]) else ""
                        col3_to = str(row[idx_3]).strip() if idx_3 < len(row) and pd.notna(row[idx_3]) else ""
                        col4_thua = str(row[idx_4]).strip() if idx_4 < len(row) and pd.notna(row[idx_4]) else ""
                        
                        if col2_name.startswith("- Vụ"):
                            current_season = col2_name
                        elif (col3_to not in ["None", "nan", ""]) and (col4_thua not in ["None", "nan", ""]):
                            parcels.append({
                                'Vụ': current_season,
                                'Tờ': col3_to,
                                'Thửa': col4_thua,
                                'Dòng Excel': excel_row_num
                            })

                    df_parcels = pd.DataFrame(parcels)
                    
                    if df_parcels.empty:
                        st.warning("⚠️ Không tìm thấy dữ liệu thửa đất nào.")
                    else:
                        parcel_counts = df_parcels.groupby(['Vụ', 'Tờ', 'Thửa']).agg(
                            Số_lần_xuất_hiện=('Tờ', 'size'),
                            Vị_trí_dòng_Excel=('Dòng Excel', lambda x: ', '.join(map(str, x))) 
                        ).reset_index()
                        parcel_counts.rename(columns={'Số_lần_xuất_hiện': 'Số lần xuất hiện', 'Vị_trí_dòng_Excel': 'Nằm tại các hàng (Excel)'}, inplace=True)

                        duplicates = parcel_counts[parcel_counts['Số lần xuất hiện'] > 1].copy()
                        
                        if not duplicates.empty:
                            st.error(f"🚨 PHÁT HIỆN THỬA ĐẤT BỊ TRÙNG LẶP TRONG CÙNG MỘT VỤ!")
                            duplicates = duplicates.reset_index(drop=True)
                            duplicates.insert(0, 'STT', range(1, len(duplicates) + 1))
                            duplicates = duplicates[['STT', 'Vụ', 'Tờ', 'Thửa', 'Số lần xuất hiện', 'Nằm tại các hàng (Excel)']]
                            st.dataframe(duplicates, use_container_width=True, hide_index=True)
                        else:
                            st.success("✅ Tuyệt vời! Không phát hiện thửa đất nào bị trùng lặp trong các vụ.")

                st.markdown("<br><h3 style='color: #1e293b; font-size: 1.2rem; margin-bottom: 10px;'>🔄 Phục hồi Data Gốc từ báo cáo PL01</h3>", unsafe_allow_html=True)
                if st.button("TRÍCH XUẤT DATA GỐC CHUẨN", type="primary", use_container_width=True):
                    with st.spinner("Đang dò tìm cột và gộp dữ liệu..."):
                        extracted_rows = []
                        current_ho_reverse = ""
                        current_season_reverse = "Không xác định"
                        current_cong_trinh_reverse = ""
                        
                        idx_1 = col_map_pl01.get('1', 0)
                        idx_2 = col_map_pl01.get('2', 1)
                        idx_3 = col_map_pl01.get('3', 2)
                        idx_4 = col_map_pl01.get('4', 3)
                        idx_5 = col_map_pl01.get('5', 4)
                        idx_6 = col_map_pl01.get('6', 5)

                        for idx, row in df_check.iloc[start_row_idx:].iterrows():
                            # BỘ LỌC FOOTER (Chống đẻ dòng ma)
                            row_str_check = " ".join([str(x).lower() for x in row.values if pd.notna(x)])
                            if "người lập" in row_str_check or "phòng kinh tế" in row_str_check or "ký tên" in row_str_check or "ngày" in str(row.iloc[17] if len(row)>17 else "").lower():
                                break
                                
                            col1_val = row[idx_1] if idx_1 < len(row) else ""
                            col1_tt = str(col1_val).strip() if pd.notna(col1_val) else ""
                            
                            col2_val = row[idx_2] if idx_2 < len(row) else ""
                            col2_name = str(col2_val).strip() if pd.notna(col2_val) else ""
                            if col2_name.lower() in ["nan", "none", "<na>"]: col2_name = ""
                            
                            col6_val = row[idx_6] if idx_6 < len(row) else ""
                            col6_name = str(col6_val).strip() if pd.notna(col6_val) else ""
                            
                            # 1. Bắt Tên Công Trình tại dòng Tổng cộng
                            if col2_name.lower() == "tổng cộng":
                                if col6_name and col6_name.lower() not in ["nan", "none", "<na>"]:
                                    current_cong_trinh_reverse = col6_name
                                continue 
                            
                            # 2. Bỏ qua các dòng Vụ tổng hợp
                            if col2_name.startswith("Vụ "):
                                continue
                            
                            # 3. Bắt Tên Hộ
                            if col1_tt.isdigit() or (col2_name != "" and not col2_name.startswith("- Vụ")):
                                current_ho_reverse = col2_name
                                
                            # 4. Bắt Tên Mùa Vụ
                            if col2_name.startswith("- Vụ"):
                                current_season_reverse = col2_name
                                
                            # 5. Extract Data Thửa Đất
                            if col2_name == "" and current_ho_reverse != "":
                                has_data = False
                                
                                target_indices = [8, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19, 20, 22, 23, 24, 25, 26, 27, 28]
                                
                                # CHỈ LẤY NẾU DIỆN TÍCH THỰC SỰ LỚN HƠN 0
                                for i in target_indices:
                                    a_idx = col_map_pl01.get(str(i+1), -1)
                                    if a_idx != -1 and a_idx < len(row):
                                        val_str = str(row[a_idx]).strip()
                                        if val_str not in ["", "nan", "None", "0", "0.0", "-"]:
                                            try:
                                                if float(val_str.replace(',', '')) > 0:
                                                    has_data = True
                                                    break
                                            except:
                                                pass
                                        
                                if has_data:
                                    new_row = [""] * 30
                                    new_row[1] = current_ho_reverse
                                    
                                    col3_val = row[idx_3] if idx_3 < len(row) else ""
                                    new_row[2] = str(col3_val).strip() if pd.notna(col3_val) and str(col3_val).strip().lower() not in ['nan','none'] else ""
                                    
                                    col4_val = row[idx_4] if idx_4 < len(row) else ""
                                    new_row[3] = str(col4_val).strip() if pd.notna(col4_val) and str(col4_val).strip().lower() not in ['nan','none'] else ""
                                    
                                    new_row[4] = row[idx_5] if idx_5 < len(row) and pd.notna(row[idx_5]) else ""
                                    new_row[5] = current_cong_trinh_reverse 
                                    
                                    for i in target_indices:
                                        actual_idx = col_map_pl01.get(str(i+1), -1)
                                        if actual_idx != -1 and actual_idx < len(row):
                                            try:
                                                new_row[i] = to_float(row[actual_idx])
                                            except:
                                                new_row[i] = 0.0
                                    
                                    new_row.append(current_season_reverse)
                                    extracted_rows.append(new_row)
                                
                        if extracted_rows:
                            cols_with_season = COLS + ['Season']
                            df_extracted = pd.DataFrame(extracted_rows, columns=cols_with_season)
                            
                            target_cols = [COLS[i] for i in target_indices]
                            for col in target_cols:
                                df_extracted[col] = pd.to_numeric(df_extracted[col], errors='coerce').fillna(0)
                            
                            df_extracted['Dup_Idx'] = df_extracted.groupby(['2', '3', '4', '5', '6', 'Season']).cumcount()
                            
                            agg_funcs = {col: 'first' for col in COLS}
                            for col in target_cols: 
                                agg_funcs[col] = 'max' 
                                
                            df_recovered = df_extracted.groupby(['2', '3', '4', '5', '6', 'Dup_Idx'], sort=False, as_index=False).agg(agg_funcs)
                            df_recovered = df_recovered[COLS] 
                            
                            recovered_excel = export_formatted_data_goc(df_recovered)
                            
                            st.success("✅ Trích xuất thành công! Dữ liệu đã được gộp chuẩn xác, tách riêng theo Công Trình.")
                            st.download_button(
                                label="📥 TẢI XUỐNG DATA GỐC ĐÃ PHỤC HỒI", 
                                data=recovered_excel, 
                                file_name="Data_Goc_Phuc_Hoi.xlsx", 
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        else:
                            st.warning("⚠️ Không tìm thấy dữ liệu hợp lệ để phục hồi. Hãy đảm bảo File PL01 đúng chuẩn.")
                        
        except Exception as e:
            st.error(f"❌ Lỗi đọc file: Vui lòng kiểm tra lại định dạng file Excel. Lỗi chi tiết: {e}")
